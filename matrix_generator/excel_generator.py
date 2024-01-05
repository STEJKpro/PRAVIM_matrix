
import requests
import pandas as pd
import os
from PIL import Image as PIL_Image

from openpyxl import Workbook, comments
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import  XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU as p2e, pixels_to_points
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.cell.text import InlineFont
import re
from io import BytesIO

from datetime import date
IMG_OFFSET = 2 #Отступ в пикселях от угла ячейки к которой приязано изображение (использовать меньше 2 не рекомендуется)

ATRR_NAMES = ['Торговая марка', 'Артикул', 'Материал', 'Цвет', 'Розничная цена (ориентировочно), BYN']


def resized_image(size,img_url):
    cached_image_path = f'cached_images/{size}/{img_url.split("/")[-1]}'
    if not os.path.exists(cached_image_path):
        os.makedirs(os.path.join(*cached_image_path.split('/')[:-1]),exist_ok=True)
        
        image = PIL_Image.open(requests.get(img_url, stream = True).raw)
        image.thumbnail((size,size), PIL_Image.LANCZOS)
        image.save(cached_image_path)
    return cached_image_path, img_url


def get_pricelist(
    urls=[
        'https://docs.google.com/spreadsheets/d/1MxIv111Z1NnP6qbcJ_QIL5FM8IWlp4qAn3ifCUGZooM/export?format=csv&id=1MxIv111Z1NnP6qbcJ_QIL5FM8IWlp4qAn3ifCUGZooM&gid=1949199208',
        'https://docs.google.com/spreadsheets/d/1MxIv111Z1NnP6qbcJ_QIL5FM8IWlp4qAn3ifCUGZooM/export?format=csv&id=1MxIv111Z1NnP6qbcJ_QIL5FM8IWlp4qAn3ifCUGZooM&gid=125708755'
    ],
    get_updates=1
):
    df = pd.DataFrame()
    if get_updates == 1:       
        for url in urls:
            response = requests.get(url)
            with open(f'matrix_generator/price/price_{urls.index(url)}.csv', 'wb') as file:
                file.write(response.content)
            df = pd.concat([df, pd.read_csv(url, encoding='utf-8', usecols=['Артикул','РРЦ с НДС, BYN (справочно)', 'Фото', 'Бренд', 'Материал', 'Цвет'], index_col='Артикул')])
        return df
    if get_updates == 0:
        for root, dirs, files in os.walk('./matrix_generator/price/'):
            for name in files:
                if name.startswith('price_'):
                    df = pd.concat([df, pd.read_csv(os.path.join(root, name), encoding='utf-8', usecols=['Артикул','РРЦ с НДС, BYN (справочно)', 'Фото', 'Бренд', 'Материал', 'Цвет'], index_col='Артикул')])
        return df
    else: raise FileNotFoundError('Файлы с прайс-листами не найдены')

def markdown_to_richtext(text):
    # Создаем пустой объект CellRichText
    richtext = CellRichText()

    # Определяем шаблоны для разных стилей markdown
    bold_pattern = r"\*\*([^\*]+)\*\*"
    # Определяем стили для разных шаблонов
    bold_style = InlineFont(b=True)
    # italic_style = InlineFont(i=True)


    # Ищем все совпадения с шаблонами в строке с markdown
    bold_matches = re.finditer(bold_pattern, text)
    # italic_matches = re.finditer(italic_pattern, markdown)

    # Создаем список из всех совпадений с их индексами и стилями
    matches = []
    for match in bold_matches:
        matches.append((match.start(), match.end(), match.group(1), bold_style))
    # for match in italic_matches:
    #     matches.append((match.start(), match.end(), match.group(1), italic_style))
    # for match in strike_matches:
    #     matches.append((match.start(), match.end(), match.group(1), strike_style))
    # Сортируем список по индексам совпадений
    matches.sort()

    # Добавляем в объект CellRichText фрагменты текста с соответствующими стилями
    start = 0
    for match in matches:
        # Добавляем обычный текст до совпадения
        if len(text[start:match[0]]) > 0:
            richtext.append(str(text[start:match[0]]))
        # Добавляем текст совпадения с нужным стилем
        richtext.append(TextBlock(text = match[2], font = match[3]))
        # Обновляем индекс начала следующего фрагмента
        start = match[1]
    # Добавляем оставшийся текст после последнего совпадения
    richtext.append( str(text[start:]))

    # Возвращаем объект CellRichText
    return richtext

 
def generate_excel_file(matrix, customer:str = None, img_size:int = None, signature_text = None,get_updates=1):
    print(signature_text)
    pricelist = get_pricelist(get_updates=get_updates)  
    for row in matrix:
        for i, item in enumerate(row):
            values = pricelist.loc[item]
            row[i] = {
                'photo': resized_image(img_url=values.get('Фото'), size=img_size) if img_size else values.get('Фото'),
                'brand': values.get('Бренд'),
                'vendor_code': item,
                'material': values.get('Материал'),
                'color': values.get('Цвет'),
                'price': values.get('РРЦ с НДС, BYN (справочно)'),                
            }

    wb = Workbook()
    ws = wb.active
    ws.sheet_view.showGridLines = False
    bd = Side(style='thin', color="000000")
    border_style = Border(left=bd, top=bd, right=bd, bottom=bd)
    alignment=Alignment(horizontal='center')
    
    
    #### Вставка шакпки с лого
    marker = AnchorMarker(col=2, colOff=p2e(IMG_OFFSET), row=1, rowOff=p2e(IMG_OFFSET))
    logo = Image ('matrix_generator/pravim_logo.png')
    h, w = logo.height, logo.width
    ws.row_dimensions[2].height = pixels_to_points(h+2*IMG_OFFSET+2)
    size = XDRPositiveSize2D(p2e(w), p2e(h))
    logo.anchor = OneCellAnchor(_from=marker, ext=size)
    ws.add_image(logo)
    
    #Вставка контрагента и даты создания в шапку
    bold_inline = InlineFont(b=True)
    value = CellRichText(
        [
            TextBlock(font = bold_inline, text = '- Подготовлено для: ') if customer else '',
            str(customer)+'\n' if customer else '',
            TextBlock(font = bold_inline, text = '- Создано: '),
            str(date.today().strftime('%d.%m.%Y')),
        ]
    )
    ws['B2'].alignment = Alignment(wrap_text=True)    
    ws['B2'] = value
    
    #Вставка текста подписи
    if signature_text:
        ws['F2'].alignment = Alignment(wrap_text=True)
        ws['F2'] = markdown_to_richtext(signature_text)
        ws.merge_cells('F2:H2')
        ws.row_dimensions[2].height = len(signature_text.split('\n'))*15 if ws.row_dimensions[2].height < len(signature_text.split('\n'))*15 else pixels_to_points(h+2*IMG_OFFSET+2)

    #### Вставка планограммы
    current_row = 3
    for matrix_row in matrix:
        current_collumn = 3
        #Ширина ячейки с наименованиями атрибутов 
        ws.column_dimensions[get_column_letter(2)].width = 38

        #Цикл по "строкам матрицы"
        for i, item in enumerate(matrix_row):
            # Вставка наименований атрибутов
            for name_i, name in enumerate(ATRR_NAMES):
                cell = ws.cell(row = current_row+name_i+1, column = 2, value= name).border = border_style
                #В артикул добавляем примечание о кликабельности
                if name == 'Артикул':
                    cell.comment = comments.Comment('Клик по артикулу откроет картинку в полном размере',author='ООО "ПРАВИМ"')
                
            
            # Цикл по атрибутам    
            for j, prop in enumerate(list(item.values())):
                if j == 0:
                    img = Image(prop[0])
                    
                    #Позиционирование изображение + добавление отступа для того чтобы были видны границы разметки
                    h, w = img.height, img.width
                    size = XDRPositiveSize2D(p2e(w), p2e(h))
                    column = current_collumn
                    row = current_row+j
                    marker = AnchorMarker(col=column-1, colOff=p2e(IMG_OFFSET), row=row-1, rowOff=p2e(IMG_OFFSET))
                    img.anchor = OneCellAnchor(_from=marker, ext=size)

                    ws.add_image(img)
                    
                    cell = ws.cell(row = current_row+j, column = current_collumn, value='')
                    cell.border = border_style
                    cell.alignment = alignment
                    
                    
                    #Задаем ширину и высоту строки с изображением (числа подобраны исходя из размеров)
                    if not ws.column_dimensions[get_column_letter(current_collumn)].width or ws.column_dimensions[get_column_letter(current_collumn)].width < (w+2*IMG_OFFSET)/7:
                        ws.column_dimensions[get_column_letter(current_collumn)].width = (w+2*IMG_OFFSET)/7
                    if not ws.row_dimensions[current_row+j].height or ws.row_dimensions[current_row+j].height < pixels_to_points(h+2*IMG_OFFSET+2):                      
                        ws.row_dimensions[current_row+j].height = pixels_to_points(h+2*IMG_OFFSET+2)
                else:   
                    cell = ws.cell(row = current_row+j, column = current_collumn, value = prop)
                    cell.alignment = alignment
                    cell.border = border_style

                    
            current_collumn +=1
        current_row = current_row + len(item) +1
        
    virtual_workbook = BytesIO()
    wb.save(virtual_workbook)
    return virtual_workbook.getbuffer()

# if __name__ == "__main__":
#     main()